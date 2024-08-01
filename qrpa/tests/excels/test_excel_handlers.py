import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def test_pivot_table():
    df = pd.read_excel('supermarket_sales.xlsx')

    # Select columns: 'Gender', 'Product line', 'Total'
    df = df[['Gender', 'Product line', 'Total']]

    # Make pivot table
    pivot_table = df.pivot_table(index='Gender', columns='Product line',
                                 values='Total', aggfunc='sum').round(0)

    # Export pivot table to Excel file
    pivot_table.to_excel('pivot_table.xlsx', 'Report', startrow=4)


def test_add_chart():
    # Read workbook and select sheet
    wb = load_workbook('pivot_table.xlsx')
    sheet = wb['Report']

    # Active rows and columns
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Instantiate a barchart
    barchart = BarChart()

    # Locate data and categories
    data = Reference(sheet,
                     min_col=min_column + 1,
                     max_col=max_column,
                     min_row=min_row,
                     max_row=max_row)  # including headers

    categories = Reference(sheet,
                           min_col=min_column,
                           max_col=min_column,
                           min_row=min_row + 1,
                           max_row=max_row)  # not including headers

    # Adding data and categories
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)

    # Make chart
    sheet.add_chart(barchart, "B12")
    barchart.title = 'Sales by Product line'
    barchart.style = 5  # choose the chart style

    # Save workbook
    wb.save('barchart.xlsx')


def test_apply_formulars():
    wb = load_workbook('barchart.xlsx')
    sheet = wb['Report']

    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Write an Excel formula with Python
    # sheet['B8'] = '=SUM(B6:B7)'
    # sheet['B8'].style = 'Currency'

    # Write multiple formulas with a for loop
    for i in range(min_column + 1, max_column + 1):  # (B, G+1)
        letter = get_column_letter(i)
        sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
        sheet[f'{letter}{max_row + 1}'].style = 'Currency'

    wb.save('report.xlsx')


def test_format_cells():
    wb = load_workbook('report.xlsx')
    sheet = wb['Report']

    # Add format
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = 'January'
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)

    wb.save('report_january.xlsx')


def test_pivot_to_report():
    # Putting together #2, #3, and #4 (input: pivot_table.xlsx + month , output: Report with barchart, formulas and format)
    month = 'february'

    # Read workbook and select sheet
    wb = load_workbook('pivot_table.xlsx')
    sheet = wb['Report']

    # Active rows and columns
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Instantiate a barchart
    barchart = BarChart()

    # Locate data and categories
    data = Reference(sheet, min_col=min_column + 1, max_col=max_column, min_row=min_row,
                     max_row=max_row)  # including headers
    categories = Reference(sheet, min_col=min_column, max_col=min_column, min_row=min_row + 1,
                           max_row=max_row)  # not including headers

    # Adding data and categories
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)

    # Make chart
    sheet.add_chart(barchart, "B12")
    barchart.title = 'Sales by Product line'
    barchart.style = 5  # choose the chart style

    # Write multiple formulas with a for loop
    for i in range(min_column + 1, max_column + 1):  # (B, G+1)
        letter = get_column_letter(i)
        sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
        sheet[f'{letter}{max_row + 1}'].style = 'Currency'

    # Add format
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)

    wb.save(f'report_{month}.xlsx')


def test_excel_exe():
    # Preparing script before we convert it to executable
    application_path = os.path.dirname(sys.executable)

    # Putting together #2, #3, and #4 (input: pivot_table.xlsx + month , output: Report with barchart, formulas and format)
    month = input('Introduce month: ')

    # Read workbook and select sheet
    input_path = os.path.join(application_path, 'pivot_table.xlsx')
    wb = load_workbook(input_path)
    sheet = wb['Report']

    # Active rows and columns
    min_column = wb.active.min_column
    max_column = wb.active.max_column
    min_row = wb.active.min_row
    max_row = wb.active.max_row

    # Instantiate a barchart
    barchart = BarChart()

    # Locate data and categories
    data = Reference(sheet,
                     min_col=min_column + 1,
                     max_col=max_column,
                     min_row=min_row,
                     max_row=max_row)  # including headers
    categories = Reference(sheet,
                           min_col=min_column,
                           max_col=min_column,
                           min_row=min_row + 1,
                           max_row=max_row)  # not including headers

    # Adding data and categories
    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(categories)

    # Make chart
    sheet.add_chart(barchart, "B12")
    barchart.title = 'Sales by Product line'
    barchart.style = 5  # choose the chart style

    # Write multiple formulas with a for loop
    for i in range(min_column + 1, max_column + 1):  # (B, G+1)
        letter = get_column_letter(i)
        sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
        sheet[f'{letter}{max_row + 1}'].style = 'Currency'

    # Add format
    sheet['A1'] = 'Sales Report'
    sheet['A2'] = month
    sheet['A1'].font = Font('Arial', bold=True, size=20)
    sheet['A2'].font = Font('Arial', bold=True, size=10)

    output_path = os.path.join(application_path, f'report_{month}.xlsx')
    wb.save(output_path)
